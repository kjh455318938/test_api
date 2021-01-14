from openpyxl import load_workbook
import unittest
from scripts.handle_config import do_config
from scripts.constants import TEST_DATAS_FILES_PATH

class Handle_excel:

    def __init__(self,filename,sheetname=None):
        self.filename=filename
        self.sheetname=sheetname

    def get_cases(self):      #获取用例
        wb=load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]
        head_date = tuple(ws.iter_rows(max_row=1, values_only=True))[0]
        one_list = []
        date = tuple(ws.iter_rows(min_row=2, values_only=True))  # 每一行的数据
        for one_tuple in date:
            one_list.append(dict(zip(head_date, one_tuple)))

        return one_list

    # def get_case1(self,row):  #获取指定某一行的用例
    #     return self.get_cases()[row-1]

    def write_result(self,row,actul,result):
        """
        同一个Workbook对象，如果将数据写入到多个表单中，那么只有最后一个表单能够写入成功
        :param row:   行数
        :param actul:   实际结果
        :param result:    返回状态
        :return:
        """
        other_wb=load_workbook(self.filename)  #指定文件
        if self.sheetname is None:
            other_ws=other_wb.active
        else:
            other_ws=other_wb[self.sheetname]
        if isinstance(row,int) and (2<=row<=other_ws.max_row):
            other_ws.cell(row=row,column=do_config.get_int('hang','value1'),value=actul)
            other_ws.cell(row=row,column=do_config.get_int("hang","value2"),value=result)
            other_wb.save(self.filename)
            other_wb.close()
        else:
            print("你输入的行号有误")

if __name__=='__main__':
    filename=TEST_DATAS_FILES_PATH

    do_excel=Handle_excel(filename,'add')
    cases=do_excel.get_cases()
    print(cases)
    # do_excel.write_result(2,"啊","去")
    pass

